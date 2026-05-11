"""Excel-Export mit openpyxl: Kopfzeile dunkelblau/weiß, deutsche Zahlenformate."""
from __future__ import annotations

from io import BytesIO
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Stilkonstanten
FARBE_KOPF = "FF1E3A8A"  # dunkelblau
SCHRIFT_WEISS_BOLD = Font(color="FFFFFFFF", bold=True, name="Calibri", size=11)
FILL_KOPF = PatternFill(start_color=FARBE_KOPF, end_color=FARBE_KOPF, fill_type="solid")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

ZAHLENFORMAT_INT = "#,##0"
ZAHLENFORMAT_DECIMAL = "#,##0.00"
ZAHLENFORMAT_EURO = '#,##0.00 "€"'


# Spalten-Definitionen pro Export-Typ -----------------------------------------
# Jeder Eintrag: (Anzeige-Spaltenname, Quell-Key, Zahlenformat oder None)
SPALTEN_DEFINITIONEN: dict[str, list[tuple[str, str, Optional[str]]]] = {
    "kunden": [
        ("Kundennummer", "kundennummer", None),
        ("Kundenname", "kundenname", None),
        ("Eier (Stück)", "eier", ZAHLENFORMAT_INT),
        ("Umsatz", "umsatz", ZAHLENFORMAT_EURO),
        ("Positionen", "positionen", ZAHLENFORMAT_INT),
        ("Letzter Kauf", "letzter_kauf", None),
    ],
    "artikel": [
        ("Artikel", "artikel_code", None),
        ("Menge", "menge", ZAHLENFORMAT_DECIMAL),
        ("Eier (Stück)", "eier", ZAHLENFORMAT_INT),
        ("Umsatz", "umsatz", ZAHLENFORMAT_EURO),
        ("Positionen", "positionen", ZAHLENFORMAT_INT),
    ],
    "ranking": [
        ("Kundennummer", "kundennummer", None),
        ("Kundenname", "kundenname", None),
        ("Eier (Stück)", "eier", ZAHLENFORMAT_INT),
        ("Umsatz", "umsatz", ZAHLENFORMAT_EURO),
    ],
    "kunde_monate": [
        ("Monat", "monat", None),
        ("Eier (Stück)", "eier", ZAHLENFORMAT_INT),
        ("Umsatz", "umsatz", ZAHLENFORMAT_EURO),
        ("Positionen", "positionen", ZAHLENFORMAT_INT),
    ],
    "artikel_monate": [
        ("Monat", "monat", None),
        ("Menge", "menge", ZAHLENFORMAT_DECIMAL),
        ("Eier (Stück)", "eier", ZAHLENFORMAT_INT),
        ("Umsatz", "umsatz", ZAHLENFORMAT_EURO),
    ],
    "jahresvergleich": [
        ("Monat", "monat", None),
        ("Eier (Jahr)", "jahr", ZAHLENFORMAT_INT),
        ("Eier (Vorjahr)", "vorjahr", ZAHLENFORMAT_INT),
        ("Differenz", "differenz", ZAHLENFORMAT_INT),
        ("Umsatz (Jahr)", "jahr_umsatz", ZAHLENFORMAT_EURO),
        ("Umsatz (Vorjahr)", "vorjahr_umsatz", ZAHLENFORMAT_EURO),
        ("Differenz €", "differenz_umsatz", ZAHLENFORMAT_EURO),
    ],
}


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)


def build_excel(typ: str, von: Optional[str], bis: Optional[str], daten: Sequence[dict],
                titel: Optional[str] = None) -> bytes:
    """Erzeugt ein .xlsx mit Titelzeile und formatierter Tabelle."""
    spalten = SPALTEN_DEFINITIONEN.get(typ)
    if spalten is None:
        raise ValueError(f"Unbekannter Export-Typ: {typ}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Auswertung"

    # Titelzeile -------------------------------------------------------------
    zeitraum = f"{von or '–'} bis {bis or '–'}" if (von or bis) else "Gesamtzeitraum"
    titel_text = titel or f"Kerba Bio-Ei GbR — {typ.capitalize()} — Zeitraum: {zeitraum}"
    ws.cell(row=1, column=1, value=titel_text).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(spalten))

    # Kopfzeile --------------------------------------------------------------
    for idx, (name, _, _) in enumerate(spalten, start=1):
        cell = ws.cell(row=3, column=idx, value=name)
        cell.font = SCHRIFT_WEISS_BOLD
        cell.fill = FILL_KOPF
        cell.alignment = ALIGN_CENTER

    # Daten ------------------------------------------------------------------
    for row_idx, eintrag in enumerate(daten, start=4):
        for col_idx, (_, key, num_fmt) in enumerate(spalten, start=1):
            wert = eintrag.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=wert)
            if num_fmt and isinstance(wert, (int, float)):
                cell.number_format = num_fmt

    ws.freeze_panes = "A4"
    _autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_excel", "SPALTEN_DEFINITIONEN"]
